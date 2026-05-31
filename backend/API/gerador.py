import io
import xml.etree.ElementTree as ET
from datetime import datetime
import openpyxl
import os

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Spacer

from backend.API.config_pdf import *
from starlette.responses import StreamingResponse


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_CLIENTE = os.path.join(BASE_DIR, "usuario", "relatorio_usuario.html")

def ajustar_colunas(aba):
    for coluna in aba.columns:
        maior = 0
        for celula in coluna:
            if celula.value:
                tamanho = len(str(celula.value))
                if tamanho > maior:
                    maior = tamanho
        letra = coluna[0].column_letter
        aba.column_dimensions[letra].width = maior + 4


def gerar_xml(dados):
    raiz = ET.Element("relatorio")

    info = ET.SubElement(raiz, "informacoes_gerais")
    ET.SubElement(info, "tipo_relatorio").text = dados["tipo_relatorio"]
    ET.SubElement(info, "nome_sistema").text    = dados["nome_sistema"]
    ET.SubElement(info, "data_geracao").text    = dados["data_geracao"]

    entidade = ET.SubElement(raiz, "entidade")
    ET.SubElement(entidade, "nome").text          = dados.get("usuario_nome")
    ET.SubElement(entidade, "email").text         = dados.get("usuario_email")
    ET.SubElement(entidade, "telefone").text      = dados.get("usuario_telefone")
    ET.SubElement(entidade, "cidade").text        = dados.get("usuario_cidade")
    ET.SubElement(entidade, "estado").text        = dados.get("usuario_estado")

    resumo = ET.SubElement(raiz, "resumo_geral")
    ET.SubElement(resumo, "total_receitas").text = str(dados["total_receitas"])
    ET.SubElement(resumo, "qtd_receitas").text = str(dados["qtd_receitas"])
    ET.SubElement(resumo, "total_despesas").text = str(dados["total_despesas"])
    ET.SubElement(resumo, "qtd_despesas").text = str(dados["qtd_despesas"])
    ET.SubElement(resumo, "saldo_periodo").text = str(dados["saldo_periodo"])
    ET.SubElement(resumo, "status_saldo").text = str(dados["status_saldo"])

    receitas = ET.SubElement(raiz, "receitas")
    for item in dados.get("receitas", []):
        receita = ET.SubElement(receitas, "receita")
        ET.SubElement(receita, "data").text = str(item.get("data_da_receita", ""))
        ET.SubElement(receita, "tipo").text = item.get("tipo_da_receita", "")
        ET.SubElement(receita, "pagador").text = item.get("pagador_email", "")
        ET.SubElement(receita, "valor").text = str(item.get("valor_da_receita", ""))
    ET.SubElement(receitas, "total").text = dados.get("total_receitas_tabela", "")

    despesas_el = ET.SubElement(raiz, "despesas")
    for item in dados.get("despesas", []):
        despesa = ET.SubElement(despesas_el, "despesa")
        data = item.get("data_da_despesa", "")
        ET.SubElement(despesa, "data").text = data.strftime("%d/%m/%Y %H:%M") if isinstance(data, datetime) else str(
            data)
        ET.SubElement(despesa, "tipo").text = item.get("tipo_da_despesa", "")
        ET.SubElement(despesa, "descricao").text = item.get("descricao_da_despesa", "")
        ET.SubElement(despesa, "recebedor").text = item.get("recebedor_email", "")
        ET.SubElement(despesa, "valor").text = str(item.get("valor_total_da_despesa", ""))
    ET.SubElement(despesas_el, "total").text = dados.get("total_despesas", "0")

    arvore = ET.ElementTree(raiz)
    ET.indent(arvore, space="    ")
    buffer = io.BytesIO()
    arvore.write(buffer, encoding="utf-8", xml_declaration=True)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/xml",
        headers={"Content-Disposition": "attachment; filename=relatorio.xml"}
    )

def gerar_xls(dados):
    wb = openpyxl.Workbook()

    aba_info = wb.active
    aba_info.title = "Informações"
    aba_info.append(["INFORMAÇÕES GERAIS"])
    aba_info.append(["Tipo de Relatório", dados["tipo_relatorio"]])
    aba_info.append(["Nome do Sistema",   dados["nome_sistema"]])
    aba_info.append(["Data de Geração",   dados["data_geracao"]])
    aba_info.append([])
    aba_info.append(["DADOS DO USUÁRIO"])
    aba_info.append(["Nome",     dados.get("usuario_nome", "")])
    aba_info.append(["E-mail",   dados.get("usuario_email", "")])
    aba_info.append(["Telefone", dados.get("usuario_telefone", "")])
    aba_info.append(["Cidade",   dados.get("usuario_cidade", "")])
    aba_info.append(["Estado",   dados.get("usuario_estado", "")])

    aba_resumo = wb.create_sheet("Resumo Geral")
    aba_resumo.append(["RESUMO GERAL"])
    aba_resumo.append(["Total de Receitas", dados["total_receitas"]])
    aba_resumo.append(["Qtd. de Receitas",  dados["qtd_receitas"]])
    aba_resumo.append(["Total de Despesas", dados["total_despesas"]])
    aba_resumo.append(["Qtd. de Despesas",  dados["qtd_despesas"]])
    aba_resumo.append(["Saldo do Período",  dados["saldo_periodo"]])
    aba_resumo.append(["Status do Saldo",   dados["status_saldo"]])

    aba_receitas = wb.create_sheet("Receitas")
    aba_receitas.append(["Nº", "Data", "Tipo", "Pagador", "Forma Pagamento", "Valor (R$)"])
    for i, item in enumerate(dados.get("receitas", []), 1):
        data = item.get("data_da_receita", "")
        aba_receitas.append([
            i,
            data.strftime("%d/%m/%Y") if hasattr(data, "strftime") else str(data),
            item.get("tipo_da_receita", ""),
            item.get("pagador_email", ""),
            item.get("forma_de_pagamento", ""),
            item.get("valor_da_receita", ""),
        ])
    aba_receitas.append(["", "", "", "", "TOTAL", dados.get("total_receitas", "")])

    aba_despesas = wb.create_sheet("Despesas")
    aba_despesas.append(["Nº", "Tipo", "Descrição", "Recebedor", "Forma Pagamento", "Valor (R$)"])
    for i, item in enumerate(dados.get("despesas", []), 1):
        aba_despesas.append([
            i,
            item.get("tipo_da_despesa", ""),
            item.get("descricao_da_despesa", ""),
            item.get("recebedor_email", ""),
            item.get("forma_de_pagamento", ""),
            item.get("valor_total_da_despesa", ""),
        ])
    aba_despesas.append(["", "", "", "", "TOTAL", dados.get("total_despesas", "")])


    ajustar_colunas(aba_info)
    ajustar_colunas(aba_resumo)
    ajustar_colunas(aba_receitas)
    ajustar_colunas(aba_despesas)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=relatorio.xlsx"}
    )


def gerar_pdf(dados):
    buffer = io.BytesIO()
    largura_util = A4[0] - 30 * mm

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=18 * mm,
        bottomMargin=20 * mm,
    )

    story = []

    # ── Cabeçalho ──────────────────────────────────────────────
    story.append(bloco_cabecalho(dados, largura_util))
    story.append(Spacer(1, 10))

    # ── Identificação ──────────────────────────────────────────
    story.append(titulo_secao("Identificação do Usuário"))
    story.append(linha_secao(largura_util))
    story.append(bloco_identificacao(dados, largura_util))
    story.append(Spacer(1, 6))

    # ── Resumo Geral ───────────────────────────────────────────
    story.append(titulo_secao("Resumo Geral"))
    story.append(linha_secao(largura_util))
    story.append(bloco_resumo(dados, largura_util))
    story.append(Spacer(1, 6))

    # ── Receitas ───────────────────────────────────────────────
    story.append(titulo_secao("Receitas"))
    story.append(linha_secao(largura_util))
    linhas_rec = []
    for r in dados.get("receitas", []):
        data = r.get("data_da_receita", "")
        if hasattr(data, "strftime"):
            data = data.strftime("%d/%m/%Y")
        linhas_rec.append([
            str(data),
            r.get("tipo_da_receita", ""),
            r.get("pagador_email", ""),
            r.get("forma_de_pagamento", ""),
            str(r.get("valor_da_receita", "")),
        ])
    story.append(tabela_dados(
        ["Data", "Tipo", "Pagador", "Forma Pagamento", "Valor (R$)"],
        linhas_rec,
        ["", "", "", "Total", dados.get("total_receitas", "")],
        largura_util,
    ))
    story.append(Spacer(1, 6))

    # ── Despesas ───────────────────────────────────────────────
    story.append(titulo_secao("Despesas"))
    story.append(linha_secao(largura_util))
    linhas_desp = []
    for d in dados.get("despesas", []):
        linhas_desp.append([
            d.get("tipo_da_despesa", ""),
            d.get("descricao_da_despesa", ""),
            d.get("recebedor_email", ""),
            d.get("forma_de_pagamento", ""),
            str(d.get("valor_total_da_despesa", "")),
        ])
    story.append(tabela_dados(
        ["Tipo", "Descrição", "Recebedor", "Forma Pagamento", "Valor (R$)"],
        linhas_desp,
        ["", "", "", "Total", dados.get("total_despesas", "")],
        largura_util,
    ))

    # ── Rodapé ─────────────────────────────────────────────────
    story.append(Spacer(1, 14))
    story.append(HRFlowable(width=largura_util, thickness=0.5,
                            color=colors.HexColor("#dddddd"), spaceAfter=4))
    est_footer = ParagraphStyle("footer", fontName="Helvetica", fontSize=8,
                                textColor=CINZA_E)
    story.append(Paragraph(
        f"Gerado em: {dados.get('data_geracao', '')} &nbsp;&nbsp;|&nbsp;&nbsp; {dados.get('nome_sistema', '')}",
        est_footer
    ))

    doc.build(story)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=relatorio.pdf"}
    )